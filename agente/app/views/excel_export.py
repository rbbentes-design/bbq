"""
Views: Excel Export

Exporta dados do bundle, curadoria, preços, risco e cenários
para um arquivo .xlsx profissional com múltiplas abas.

Inspirado no Claude Office Skills (tfriedel/claude-office-skills).
Requer: pip install openpyxl

Abas geradas:
  1. Resumo        — visão geral do run
  2. Preços        — cotações e retornos de mercado
  3. Técnicos      — RSI, MACD, Bollinger
  4. Risco         — VaR, CVaR, MaxDD, Sharpe
  5. Cenários      — Bull / Base / Bear
  6. ISQ           — cadeia causal e tickers impactados
  7. Polymarket    — probabilidades de mercado
  8. Monte Carlo   — distribuição final por ticker

Uso:
    from app.views.excel_export import export_to_excel
    path = export_to_excel(bundle, curation_result, enrichment, output_path)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from app.audit.logger import get_logger

_log = get_logger("views.excel_export")


def export_to_excel(
    bundle: Any,
    curation_result: Any | None,
    enrichment: dict[str, Any],
    output_path: Path,
) -> Path | None:
    """
    Exporta todos os dados para Excel (.xlsx).

    Args:
        bundle:          DailyIngestionBundle
        curation_result: CurationResult ou None
        enrichment:      dict com chaves: market_prices, technical, risk,
                         scenarios, isq_signal, polymarket, monte_carlo
        output_path:     Caminho de saída .xlsx

    Returns:
        Path do arquivo gerado, ou None se falhar.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        _log.warning("openpyxl_not_installed", hint="pip install openpyxl")
        return None

    wb = Workbook()
    wb.remove(wb.active)  # remove aba default

    # Estilos reutilizáveis
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
    POSITIVE_FILL = PatternFill("solid", fgColor="C6EFCE")
    NEGATIVE_FILL = PatternFill("solid", fgColor="FFC7CE")
    NEUTRAL_FILL = PatternFill("solid", fgColor="FFEB9C")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    WRAP = Alignment(wrap_text=True, vertical="top")

    def _header(ws: Any, row: int, cols: list[str]) -> None:
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    def _color_pct(val: float | None) -> PatternFill | None:
        if val is None:
            return None
        if val > 0.001:
            return POSITIVE_FILL
        if val < -0.001:
            return NEGATIVE_FILL
        return NEUTRAL_FILL

    def _autowidth(ws: Any, min_width: int = 8, max_width: int = 40) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    # ── 1. Resumo ──────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Resumo")
    ws1.freeze_panes = "A2"
    summary_rows = [
        ("Run ID", bundle.run_id),
        ("Data", str(bundle.run_date)),
        ("Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("ZeroHedge Blocos", len(bundle.market_ear_blocks)),
        ("X Items", len(bundle.x_items)),
        ("RSS Items", len(bundle.rss_items)),
        ("SpotGamma Reports", len(bundle.spotgamma_reports)),
    ]
    if curation_result:
        primary = curation_result.narrative.primary_signal
        summary_rows += [
            ("", ""),
            ("=== CURADORIA ===", ""),
            ("Narrativa Primária", primary.label),
            ("Confiança", f"{primary.confidence:.0%}"),
            ("Verificação", curation_result.verification.overall_verdict),
            ("Itens Pontuados", len(curation_result.scored_items)),
        ]
    for r, (k, v) in enumerate(summary_rows, 1):
        ws1.cell(r, 1, k).font = Font(bold=True)
        ws1.cell(r, 2, v)
    _autowidth(ws1)

    # ── 2. Preços ──────────────────────────────────────────────────────────────
    market_prices: dict[str, Any] = enrichment.get("market_prices", {})
    if market_prices:
        ws2 = wb.create_sheet("Preços")
        ws2.freeze_panes = "A2"
        _header(ws2, 1, ["Símbolo", "Nome", "Preço", "Retorno Diário", "Retorno Semanal", "YTD"])
        for r, (sym, d) in enumerate(market_prices.items(), 2):
            ws2.cell(r, 1, sym)
            ws2.cell(r, 2, d.get("name", sym))
            ws2.cell(r, 3, d.get("price"))
            for col_idx, key in [(4, "daily_return"), (5, "weekly_return"), (6, "ytd_return")]:
                val = d.get(key)
                cell = ws2.cell(r, col_idx, f"{val:+.2%}" if val is not None else "N/A")
                fill = _color_pct(val)
                if fill:
                    cell.fill = fill
                cell.alignment = CENTER
        _autowidth(ws2)

    # ── 3. Técnicos ────────────────────────────────────────────────────────────
    technical: dict[str, Any] = enrichment.get("technical", {})
    if technical:
        ws3 = wb.create_sheet("Técnicos")
        ws3.freeze_panes = "A2"
        _header(ws3, 1, ["Símbolo", "Nome", "RSI(14)", "MACD", "Signal Line",
                          "Histograma", "BB Upper", "BB Middle", "BB Lower", "%B", "Sinal"])
        for r, (sym, ind) in enumerate(technical.items(), 2):
            name = market_prices.get(sym, {}).get("name", sym)
            m = ind.get("macd", {})
            bb = ind.get("bollinger", {})
            ws3.cell(r, 1, sym)
            ws3.cell(r, 2, name)
            rsi_val = ind.get("rsi")
            rsi_cell = ws3.cell(r, 3, rsi_val)
            if rsi_val is not None:
                if rsi_val < 30:
                    rsi_cell.fill = POSITIVE_FILL
                elif rsi_val > 70:
                    rsi_cell.fill = NEGATIVE_FILL
            for col_idx, val in [
                (4, m.get("macd_line")), (5, m.get("signal_line")), (6, m.get("histogram")),
                (7, bb.get("upper")),   (8, bb.get("middle")),       (9, bb.get("lower")),
                (10, bb.get("pct_b")),
            ]:
                ws3.cell(r, col_idx, val)
            ws3.cell(r, 11, ind.get("signal", ""))
        _autowidth(ws3)

    # ── 4. Risco ───────────────────────────────────────────────────────────────
    risk: dict[str, Any] = enrichment.get("risk", {})
    tickers_risk = risk.get("tickers", {})
    if tickers_risk:
        ws4 = wb.create_sheet("Risco")
        ws4.freeze_panes = "A2"
        _header(ws4, 1, ["Símbolo", "Nome", "VaR 95%", "CVaR 95%", "Max Drawdown", "Sharpe"])
        for r, (sym, m) in enumerate(tickers_risk.items(), 2):
            name = market_prices.get(sym, {}).get("name", sym)
            ws4.cell(r, 1, sym)
            ws4.cell(r, 2, name)
            for col_idx, key in [(3, "var_95"), (4, "cvar_95"), (5, "max_drawdown")]:
                val = m.get(key)
                cell = ws4.cell(r, col_idx, f"{val:+.2%}" if val is not None else "N/A")
                if val is not None and val < -0.02:
                    cell.fill = NEGATIVE_FILL
                cell.alignment = CENTER
            sharpe = m.get("sharpe")
            cell = ws4.cell(r, 6, sharpe)
            if sharpe is not None:
                cell.fill = POSITIVE_FILL if sharpe > 0.5 else (NEGATIVE_FILL if sharpe < 0 else NEUTRAL_FILL)
            cell.alignment = CENTER
        _autowidth(ws4)

    # ── 5. Cenários ────────────────────────────────────────────────────────────
    scenarios: dict[str, Any] = enrichment.get("scenarios", {})
    if scenarios:
        ws5 = wb.create_sheet("Cenários")
        ws5.freeze_panes = "A2"
        _header(ws5, 1, ["Cenário", "Probabilidade", "Catalisador", "Narrativa", "SPX Target", "Horizonte"])
        fills = {"bull": POSITIVE_FILL, "base": NEUTRAL_FILL, "bear": NEGATIVE_FILL}
        for r, label in enumerate(["bull", "base", "bear"], 2):
            s = scenarios.get(label, {})
            ws5.cell(r, 1, label.upper()).fill = fills.get(label, NEUTRAL_FILL)
            ws5.cell(r, 2, f"{s.get('probability', 0):.0%}").alignment = CENTER
            ws5.cell(r, 3, s.get("catalyst", ""))
            ws5.cell(r, 4, s.get("narrative", "")).alignment = WRAP
            ws5.cell(r, 5, s.get("spx_target", "")).alignment = CENTER
            ws5.cell(r, 6, s.get("time_horizon", ""))
        ws5.row_dimensions[2].height = 50
        ws5.row_dimensions[3].height = 50
        ws5.row_dimensions[4].height = 50
        _autowidth(ws5, max_width=60)

    # ── 6. ISQ ─────────────────────────────────────────────────────────────────
    isq: Any = enrichment.get("isq_signal")
    if isq:
        ws6 = wb.create_sheet("ISQ")
        ws6.cell(1, 1, "INVESTMENT SIGNAL QUALIFICATION").font = Font(bold=True, size=12)
        ws6.cell(2, 1, isq.title)
        ws6.cell(3, 1, f"Sentimento: {isq.sentiment_score:+.2f} | Confiança: {isq.confidence:.0%} | Intensidade: {isq.intensity}/5")

        ws6.cell(5, 1, "CADEIA DE TRANSMISSÃO").font = Font(bold=True)
        _header(ws6, 6, ["#", "Nó", "Impacto", "Justificativa"])
        for i, node in enumerate(isq.transmission_chain, 1):
            r = 6 + i
            ws6.cell(r, 1, i)
            ws6.cell(r, 2, node.node)
            color = {"positive": "00B050", "negative": "FF0000", "neutral": "7F7F7F"}.get(node.impact, "7F7F7F")
            ws6.cell(r, 3, node.impact.upper()).font = Font(color=color, bold=True)
            ws6.cell(r, 4, node.reasoning).alignment = WRAP
            ws6.row_dimensions[r].height = 35

        offset = 6 + len(isq.transmission_chain) + 2
        ws6.cell(offset, 1, "ATIVOS IMPACTADOS").font = Font(bold=True)
        _header(ws6, offset + 1, ["Ticker", "Nome", "Direção", "Peso", "Justificativa"])
        for i, ticker in enumerate(isq.impact_tickers, 1):
            r = offset + 1 + i
            ws6.cell(r, 1, ticker.ticker)
            ws6.cell(r, 2, ticker.name or ticker.ticker)
            dir_colors = {"long": "00B050", "short": "FF0000", "neutral": "7F7F7F"}
            ws6.cell(r, 3, ticker.direction.upper()).font = Font(
                color=dir_colors.get(ticker.direction, "7F7F7F"), bold=True
            )
            ws6.cell(r, 4, f"{ticker.weight:.0%}").alignment = CENTER
            ws6.cell(r, 5, ticker.reasoning).alignment = WRAP
            ws6.row_dimensions[r].height = 35
        _autowidth(ws6, max_width=60)

    # ── 7. Polymarket ──────────────────────────────────────────────────────────
    polymarket: list[dict] = enrichment.get("polymarket", [])
    if polymarket:
        ws7 = wb.create_sheet("Polymarket")
        ws7.freeze_panes = "A2"
        _header(ws7, 1, ["Questão", "Probabilidade (Yes)", "Volume (USD)", "Vencimento"])
        for r, m in enumerate(polymarket, 2):
            ws7.cell(r, 1, m.get("question", ""))
            prob = m.get("probability", 0)
            cell = ws7.cell(r, 2, f"{prob:.1%}")
            cell.alignment = CENTER
            cell.fill = POSITIVE_FILL if prob > 0.6 else (NEGATIVE_FILL if prob < 0.3 else NEUTRAL_FILL)
            ws7.cell(r, 3, f"${m.get('volume_usd', 0):,.0f}").alignment = CENTER
            ws7.cell(r, 4, m.get("end_date", ""))
        _autowidth(ws7)

    # ── 8. Monte Carlo ─────────────────────────────────────────────────────────
    monte_carlo: dict[str, Any] = enrichment.get("monte_carlo", {})
    if monte_carlo:
        ws8 = wb.create_sheet("Monte Carlo")
        ws8.freeze_panes = "A2"
        _header(ws8, 1, ["Símbolo", "Nome", "Preço Atual", "Horizonte",
                          "P5", "Média", "P95", "P(Alta)", "P(+5%)", "P(-5%)"])
        for r, (sym, mc) in enumerate(monte_carlo.items(), 2):
            name = market_prices.get(sym, {}).get("name", sym)
            fd = mc.get("final_distribution", {})
            ws8.cell(r, 1, sym)
            ws8.cell(r, 2, name)
            ws8.cell(r, 3, mc.get("current_price"))
            ws8.cell(r, 4, f"{mc.get('horizon_days', 0)}d").alignment = CENTER
            ws8.cell(r, 5, fd.get("p5"))
            ws8.cell(r, 6, fd.get("mean"))
            ws8.cell(r, 7, fd.get("p95"))
            for col_idx, key in [(8, "prob_up"), (9, "prob_up_5pct"), (10, "prob_down_5pct")]:
                val = mc.get(key)
                cell = ws8.cell(r, col_idx, f"{val:.0%}" if val is not None else "N/A")
                if val is not None:
                    cell.fill = POSITIVE_FILL if (col_idx < 10 and val > 0.5) else (
                        NEGATIVE_FILL if (col_idx == 10 and val > 0.3) else NEUTRAL_FILL
                    )
                cell.alignment = CENTER
        _autowidth(ws8)

    # ── Salva ──────────────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    _log.info("excel_saved", path=str(output_path), sheets=len(wb.sheetnames))
    return output_path
